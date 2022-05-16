import pyttsx3
import openpyxl
#import phonetics
#import eng_to_ipa as p
import re
import speech_recognition as sr

engine=pyttsx3.init()

wrkbk=openpyxl.load_workbook("Input File.xlsx")

sheet=wrkbk.active

for rowNumber in range(2, sheet.max_row + 1):
    #text=phonetics.metaphone(sheet.cell(row=rowNumber, column=4).value)
    text=sheet.cell(row=rowNumber, column=4).value
    if sheet.cell(row=rowNumber, column=5).value == 'No':
        #text=p.convert(sheet.cell(row=rowNumber, column=4).value)
        #print(re.sub('\W+','',text))
        filename=(sheet.cell(row=rowNumber, column=4).value+'.m4a')
        engine.save_to_file(re.sub('\W+','',text),filename)
        engine.runAndWait()
    else:
        r = sr.Recognizer()
        filename=(sheet.cell(row=rowNumber, column=4).value+'.m4a')
        hellow=sr.AudioFile(filename)
        with hellow as source:
            audio = r.record(source)
            try:
                s = r.recognize_google(audio)
                filename=(sheet.cell(row=rowNumber, column=4).value+'.m4a')
                engine.save_to_file(re.sub('\W+','',s),filename)
                engine.runAndWait()
            except Exception as e:
                print("Exception: "+str(e))